#!/usr/bin/env python
# encoding=utf-8
# from openpyxl import Workbook
import sys    
reload(sys)  
sys.setdefaultencoding('utf8')
from openpyxl import load_workbook
keys = ['uniacid', 'openid', 'realname', 'mobile']
def get_data():
	wb = load_workbook(filename="电影.xlsx")
	ws = wb.get_sheet_names()
	print wb.sheetnames[0].encode('utf-8')
	for sheetname in ws:
		# print sheetname.isdigit()
		# if not sheetname.isdigit():
		# 	continue
		sheet = wb[sheetname]
		# print sheet.max_row
		# print sheet.max_column
		# print type(sheet.rows)
		# for x in sheet.rows:
		# 	print x

	# print type(sheet.values)
	# print type(sheet.rows)

	#按行读取
	for i,j in enumerate(sheet.values):
		# print i,j
		movie = ''
		for jid,cell in enumerate(j):
			if cell:
				if jid ==0:
					movie+='电影名：'+cell+','
				if jid ==2:
					movie+='评分：'+cell+','
				if jid ==3:
					movie+='短评：'+cell+''
		# print movie
	# for idx,row in enumerate(sheet.rows):
	# 	if idx == 0:  continue;
	# 	# print idx
	# 	# print row
	# 	vals =[]
	# 	for cell in row:
	# 		vals.append(cell.value) 
	# 	d = dict(zip(keys,vals))
	# 	print d
	# 	rs.append(d)
	# print ws.max_columb.
	# print ws.get_highest_row()
	# print ws.get_highest_column()
	# print ws.columns

	#按列读取
	for i,j in enumerate(sheet.columns):
		# print j
		for jid,cell in enumerate(j):
			if cell.value:
				print cell.value

get_data()